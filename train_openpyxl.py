"""openpyxl 모듈 실습

공식 문서 : https://openpyxl.readthedocs.io/en/stable/index.html
"""

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection

import os.path
import datetime

xl = openpyxl.Workbook()    # 엑셀파일 생성(추후 xl.save()를 해야 엑셀파일이 생성됨)

ws = xl.active   # 활성화되어있는 시트 객체 가져오기(처음엔 시트가 1개('Sheet')밖에 없어서 명칭이 'Sheet'인 시트를 가져옴)
print(ws.title)  # 시트명 출력 : Sheet

xl.create_sheet('Sheet2')   # 신규 시트 생성(생성했다고 활성화되진 않음!)
print(ws.title)  # 시트명 출력 : Sheet

ws = xl['Sheet2']    # 이렇게 특정 시트 객체를 가져올 수 있는데 이러면 ws. 할 때 관련 메소드가 안나옴... 근데 안나와도 ws.title 이런거 하면 오류는 안남
print(ws.title)  # 시트명 출력 : Sheet2

xl._active_sheet_index = 1  # 이렇게 쓰면 안될 것 같긴한데.. 잘 작동은 함
ws = xl.active
print(ws)    # <Worksheet "Sheet2">
ws.sheet_properties.tabColor = '1072BA'     # tab색상 변경, RRGGBB 형태

ws.cell(1, 1, 'abc') # 1행 1열에 abc 문자열 쓰기(row, col은 1부터 시작함)
ws['B1'] = 'B1'
ws.append([1,2,3,4,5])   # 데이터가 있는 마지막 행 다음 행에 1열에 1, 2열에 2 3열에 3, ... 쓰기

cell_range = ws['A1':'C1']
for i in cell_range:
    print(i)    # (<Cell 'Sheet2'.A1>, <Cell 'Sheet2'.B1>, <Cell 'Sheet2'.C1>)
    for cell in i:
        cell.value = datetime.datetime.now()   # cell 객체에는 value로 값을 쓸 수 있음


ws['F2'] = "=SUM(A2:E2)"    # 수식 입력


ws.merge_cells('A3:C3') # 머지
# ws.unmerge_cells('A3:D3')   # 머지해제
# ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
# ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)


img = openpyxl.drawing.image.Image('python-logo.png')   # Pillow가 내장되어있지 않으면 에러나네..
ws.add_image(img, 'B4') # 이미지 넣기

ws.insert_rows(2)   # 위치에 행 삽입
# ws.delete_rows(2)   # 해당 행 삭제
ws.insert_cols(2)   # 해당 열 삽입
# ws.delete_cols(6, 3)    # 해당 컬럼(F:H, F열(6)부터 3개열) 삭제


# ws['I5'] = 'I5'
# ws['I6'] = 'I6'
# ws['J5'] = 'J5'
# ws['J6'] = 'J6'
for i in ['I', 'J']:
    for j in ['5', '6']:
        ws[i+j] = i+j

ws.move_range("I5:J6", rows=-2, cols=2)    # 특정 범위의 셀들을 rows, cols값 만큼 이동(rows, cols '로' 이동이 아니라 '만큼' 이동)


comment = openpyxl.comments.Comment('주석', '김완영')
ws['A1'].comment = comment  # 주석


ws['A1'].font = Font(name='Calibri', sz=15, bold=True, italic=True, color=colors.RED)   # 폰트 속성값 설정 color = RRGGBB 형식 가능

# 다음 네 줄은 이미 있던 데이터에는 영향을 끼지치 않고, 신규로 입력된 데이터에 대해서만 적용됨
# 따라서 이미 있던 데이터에도 적용하고 싶다면, 각 셀별로 스타일을 적용하면 됨
col = ws.column_dimensions['K'] # K열 전체
col.font = Font(color=colors.BLUE)  # A열 전체에 대해 폰트 속성값 설정
row = ws.row_dimensions[3]  # 3행 전체
row.font = Font(color=colors.YELLOW) # 3행 전체에 폰트 속성값 설정



# 테이블 적용
data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears',   2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges',  500,  300,  200,  700],
]

# add column headings. NB. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref="A7:E11")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)


# 엑셀파일 암호 거는 법은 뭔가 안되는 듯?
# hashed_password = '123'
# xl.security.set_workbook_password(hashed_password, already_hashed=True)


# 엑셀 시트 보호
ws = xl.active
# ws.protection.sheet = True  # 시트보호를 쓸거냐 말거냐인가?
# ws.protection.enable()  # 시트보호 설정
# ws.protection.disable() # 시트보호 해제

# ws.protection.password = '123'    # 시트보호 설정하고 해제할때 필요한 암호/ 위의 스크립트 안쓰고 이 스크립트 만으로도 시트보호 설정 가능


# if not os.path.exists('openpyxl_test.xlsx'):
xl.save(filename='openpyxl_test.xlsx')   # 저장(주의:이미 있는 경우 덮어써지므로 존재여부 확인 후 저장하기)


# if os.path.exists('openpyxl_test.xlsx'):
#     xl = openpyxl.load_workbook('openpyxl_test.xlsx')   # 엑셀 로드


xl.close()
